import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import datetime
import os
import logging
import re

# Configure Logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# Calculate template path relative to this file
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
TEMPLATE_PATH = os.path.join(BASE_DIR, "templates", "solar_template.xlsx")

# Styling Tokens
THIN_BORDER = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC')
)

SECTION_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

def write_section_header(ws, row_idx, text):
    """
    Robustly formats a merged section header across A:D.
    Prevents text cutoff in Google Sheets and Apple Numbers.
    """
    # 1. Expand Merge Width to Full A:D
    cell_range = f"A{row_idx}:D{row_idx}"
    if cell_range in ws.merged_cells:
        ws.unmerge_cells(cell_range)
    ws.merge_cells(cell_range)
    
    cell = ws.cell(row=row_idx, column=1)
    cell.value = text
    
    # 2. Force Alignment & Wrap Text
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    # 3. Apply Professional Font & Background
    cell.font = Font(bold=True, size=12, color="333333")
    cell.fill = SECTION_FILL
    
    # 4. Force Row Height (MANDATORY for merged cells)
    # 40 provides enough room for wrapping long titles
    ws.row_dimensions[row_idx].height = 40
    
    # Apply borders to the full range
    for col in range(1, 5):
        ws.cell(row=row_idx, column=col).border = THIN_BORDER

def setup_page_layout(ws):
    """
    Sets initial page layout.
    """
    # Freeze header row 1
    ws.freeze_panes = 'A2'

    # Main Header A1:D1
    ws.row_dimensions[1].height = 45
    ws.merge_cells('A1:D1')
    header = ws['A1']
    header.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    header.font = Font(bold=True, size=16)

def write_data_row(ws, row_idx, label, value, number_format=None):
    """
    Writes a balanced 4-column data row (A:B merged | C:D merged).
    """
    # Label (A:B)
    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=2)
    c_label = ws.cell(row=row_idx, column=1)
    c_label.value = label
    c_label.font = Font(bold=True)
    c_label.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    # Value (C:D)
    ws.merge_cells(start_row=row_idx, start_column=3, end_row=row_idx, end_column=4)
    c_val = ws.cell(row=row_idx, column=3)
    c_val.value = value
    if number_format:
        c_val.number_format = number_format
    c_val.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.row_dimensions[row_idx].height = 25
    for col in range(1, 5):
        ws.cell(row=row_idx, column=col).border = THIN_BORDER

def clean_numeric(value):
    """Safe float conversion."""
    if value is None or value == "": return 0.0
    if isinstance(value, (int, float)): return float(value)
    cleaned = re.sub(r"[^\d.]", "", str(value).replace(',', ''))
    try: return float(cleaned)
    except: return 0.0

def apply_auto_formatting(ws):
    """
    Applies auto-adjustments to column widths, row heights, and ensures text wrapping.
    """
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    col_widths = {}
    
    for row in ws.iter_rows():
        row_max_height = 15
        
        for cell in row:
            if cell.alignment:
                cell.alignment = Alignment(
                    horizontal=cell.alignment.horizontal,
                    vertical="center",
                    wrap_text=True
                )
            else:
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                
            if cell.value is not None:
                text = str(cell.value)
                
                is_merged = False
                for merge_range in ws.merged_cells.ranges:
                    if cell.coordinate in merge_range:
                        is_merged = True
                        break
                
                if not is_merged:
                    length = len(text)
                    col_letter = cell.column_letter
                    if col_letter not in col_widths or length > col_widths[col_letter]:
                        col_widths[col_letter] = length

                num_newlines = text.count('\n')
                lines = num_newlines + 1
                
                estimated_height = lines * 15
                if estimated_height > row_max_height:
                    row_max_height = estimated_height
        
        current_height = ws.row_dimensions[row[0].row].height
        if current_height is None or current_height < row_max_height + 5:
            ws.row_dimensions[row[0].row].height = row_max_height + 5

    for col_letter, max_len in col_widths.items():
        adjusted_width = (max_len + 2) * 1.2
        if adjusted_width > 50:
            adjusted_width = 50
        if adjusted_width < 15:
            adjusted_width = 15
        ws.column_dimensions[col_letter].width = adjusted_width

def write_to_template(payload, output_path):
    """
    Refined Excel Generation Engine with Cutoff-Prevention Logic.
    """
    data = payload.get("data", payload)
    
    if not os.path.exists(TEMPLATE_PATH):
        raise FileNotFoundError(f"Template not found at {TEMPLATE_PATH}")

    try:
        wb = openpyxl.load_workbook(TEMPLATE_PATH)
        ws = wb.active
        
        setup_page_layout(ws)
        
        # Section 1: Extraction Results
        write_section_header(ws, 2, "AI Extraction Results")
        write_data_row(ws, 3, "Consumer Name", str(data.get("consumer_name") or "Valued Customer"))
        write_data_row(ws, 4, "Consumer Number", str(data.get("consumer_number") or "N/A"))
        write_data_row(ws, 5, "Units (kWh)", int(clean_numeric(data.get("units"))), number_format='#,##0')
        write_data_row(ws, 6, "Total Bill Amount", clean_numeric(data.get("amount")), number_format='₹#,##0.00')
        write_data_row(ws, 7, "Rate (₹/kWh)", clean_numeric(data.get("tariff")), number_format='0.00')
        write_data_row(ws, 8, "Extraction Date", datetime.datetime.now().strftime("%Y-%m-%d %H:%M"))

        # Consistent spacing between sections
        ws.row_dimensions[9].height = 15

        # Section 2: Load Recommendation (Fixing the Cutoff Issue)
        write_section_header(ws, 10, "Calculated Solar Load Recommendation & Estimate")
        
        # Dynamic Load Calculation Row (Business logic preserved)
        units = clean_numeric(data.get("units"))
        solar_load = round(units / 120, 2) if units > 0 else 0
        write_data_row(ws, 11, "Recommended Solar Capacity", f"{solar_load} kW")

        # Apply formatting to both sheets
        target_sheets = ["Bill Data", "Solar Calculation"]
        for sheet_name in target_sheets:
            if sheet_name in wb.sheetnames:
                apply_auto_formatting(wb[sheet_name])
            elif sheet_name == "Bill Data" and "Sheet" in wb.sheetnames:
                wb["Sheet"].title = "Bill Data"
                apply_auto_formatting(wb["Bill Data"])
        
        if wb.active.title not in target_sheets:
            apply_auto_formatting(wb.active)

        wb.save(output_path)
        logger.info(f"Report generated successfully with cutoff-prevention: {output_path}")
        return output_path

    except Exception as e:
        logger.error(f"Excel Cutoff-Prevention Error: {str(e)}")
        raise
