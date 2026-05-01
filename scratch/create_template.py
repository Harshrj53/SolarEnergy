import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

def create_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Solar Load Calculator"

    # Header
    ws.merge_cells('A1:D1')
    ws['A1'] = "Electricity Bill Data Input"
    ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
    ws['A1'].alignment = Alignment(horizontal="center")
    ws['A1'].fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

    # Labels
    labels = [
        ("A3", "Consumer Name:"),
        ("A4", "Consumer Number:"),
        ("A5", "Units Consumed (kWh):"),
        ("A6", "Total Bill Amount (₹):"),
        ("A7", "Tariff / Rate (₹/kWh):"),
        ("A8", "Extraction Date:")
    ]

    for cell, text in labels:
        ws[cell] = text
        ws[cell].font = Font(bold=True)

    # Input Placeholders (B3:B8)
    # Formulas (Optional for demo)
    ws['A10'] = "Calculated Solar Load Recommendation"
    ws['A10'].font = Font(bold=True)
    ws['A11'] = "Estimated Panels Needed:"
    ws['B11'] = "=B5/120" # Dummy formula: Units / 120 (rough solar generation)

    wb.save("/Users/harshraj/Desktop/EnergyBae/templates/solar_template.xlsx")

if __name__ == "__main__":
    import os
    os.makedirs("/Users/harshraj/Desktop/EnergyBae/templates", exist_ok=True)
    create_template()
